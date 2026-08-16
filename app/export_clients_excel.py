"""
Export Excel des clients avec les donnees disponibles gratuitement via les
API Google Business Profile deja utilisees dans la plateforme (fiche, avis,
photos, posts) - volontairement AUCUNE donnee payante (positions/DataForSEO).
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from . import google_business, google_location, google_oauth, google_reviews

ENTETES = [
    "Nom", "Prénom", "Email", "Compte Google", "Étiquettes",
    "Téléphone", "Site web", "Adresse", "Catégorie principale",
    "Complétude fiche", "Note moyenne", "Nombre d'avis",
    "Photos sur la fiche", "Posts publiés (plateforme)",
    "Lien fiche Google", "Erreur",
]


def _adresse_lisible(infos: dict) -> str:
    adresse = infos.get("storefrontAddress") or {}
    lignes = ", ".join(adresse.get("addressLines") or [])
    reste = " ".join(p for p in [adresse.get("postalCode", ""), adresse.get("locality", "")] if p)
    return ", ".join(p for p in [lignes, reste] if p)


def generer_export(db, clients: list) -> bytes:
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Clients"
    feuille.append(ENTETES)
    for cellule in feuille[1]:
        cellule.font = Font(bold=True)

    identifiants_par_compte = {}

    for client in clients:
        nb_posts = sum(1 for p in client.posts if p.statut == "PUBLIE_LIVE")
        ligne = [
            client.nom,
            client.prenom,
            client.email,
            client.compte_google.libelle if client.compte_google else "",
            ", ".join(e.nom for e in client.etiquettes),
            "", "", "", "", "", "", "", "",
            nb_posts,
            "", "",
        ]

        if client.account_id and client.location_id:
            if client.compte_google_id not in identifiants_par_compte:
                identifiants_par_compte[client.compte_google_id] = google_oauth.obtenir_identifiants(
                    db, client.compte_google_id
                )
            identifiants = identifiants_par_compte[client.compte_google_id]

            if not identifiants:
                ligne[-1] = "Compte Google non valide"
            else:
                erreurs = []
                try:
                    infos = google_location.obtenir_infos_fiche(identifiants, client.location_id)
                    completude = google_location.score_completude(infos)
                    categorie_principale = ((infos.get("categories") or {}).get("primaryCategory") or {}).get(
                        "displayName", ""
                    )
                    ligne[5] = (infos.get("phoneNumbers") or {}).get("primaryPhone", "")
                    ligne[6] = infos.get("websiteUri", "")
                    ligne[7] = _adresse_lisible(infos)
                    ligne[8] = categorie_principale
                    ligne[9] = f"{completude['score']}/{completude['total']}"
                    ligne[14] = (infos.get("metadata") or {}).get("mapsUri", "")
                except Exception as erreur:
                    erreurs.append(f"Fiche : {erreur}")

                try:
                    resume_avis = google_reviews.resume_rapide(identifiants, client.account_id, client.location_id)
                    ligne[10] = resume_avis["note_moyenne"] if resume_avis["note_moyenne"] is not None else ""
                    ligne[11] = resume_avis["total_avis"]
                except Exception as erreur:
                    erreurs.append(f"Avis : {erreur}")

                try:
                    photos = google_business.lister_photos(identifiants, client.account_id, client.location_id)
                    ligne[12] = len(photos)
                except Exception as erreur:
                    erreurs.append(f"Photos : {erreur}")

                ligne[-1] = " ; ".join(erreurs)

        feuille.append(ligne)

    for indice, entete in enumerate(ENTETES, start=1):
        feuille.column_dimensions[get_column_letter(indice)].width = max(14, len(entete) + 2)
    feuille.freeze_panes = "A2"

    tampon = BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()
